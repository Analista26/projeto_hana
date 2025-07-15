import win32com.client as win32
import openpyxl
import os
from datetime import datetime
from tqdm import tqdm  # progress bar

def preencher_todas_abas(template_path, dados_path, output_path, row_start_dest=9):
    """
    Preenche todas as abas do template SAP Migration Cockpit (LTMC/LTMOM) com dados de um XLSX,
    desde que existam abas com o mesmo nome em ambos.
    """

    log = []
    inicio = datetime.now()
    log.append(f"[{inicio}] InÃ­cio processamento: {os.path.basename(template_path)}")

    # Carrega XLSX origem
    wb_origem = openpyxl.load_workbook(dados_path, data_only=True)

    # Inicia Excel COM
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    wb_template = excel.Workbooks.Open(template_path)

    abas_xml = [ws.Name for ws in wb_template.Worksheets]
    log.append(f"Abas disponÃ­veis no template XML: {abas_xml}")

    # Loop por todas as abas do XLSX
    for aba_origem in wb_origem.sheetnames:
        ws_origem = wb_origem[aba_origem]

        # Verifica se a aba existe no XML
        if aba_origem not in abas_xml:
            log.append(f"âŒ Aba '{aba_origem}' nÃ£o encontrada no template XML. Pulando...")
            continue

        # Coleta dados a partir da linha definida
        rows = list(ws_origem.iter_rows(min_row=row_start_dest, values_only=True))
        total_rows = len(rows)

        # Verifica se hÃ¡ dados relevantes
        if not any(any(cell is not None for cell in row) for row in rows):
            log.append(f"ðŸ”¶ Aba '{aba_origem}' ignorada (sem dados na linha {row_start_dest}).")
            continue

        log.append(f"ðŸ”µ Preenchendo aba '{aba_origem}' com {total_rows} linhas...")

        ws_template = wb_template.Worksheets(aba_origem)
        ws_template.Unprotect()

        row_dest = row_start_dest

        for row in tqdm(rows, desc=f"Preenchendo {aba_origem}", unit="linha"):
            col_dest = 1
            for value in row:
                ws_template.Cells(row_dest, col_dest).Value = value
                col_dest += 1
            row_dest += 1

    # Salva como XML Spreadsheet 2003
    wb_template.SaveAs(output_path, FileFormat=46)
    wb_template.Close(SaveChanges=True)
    excel.Quit()

    fim = datetime.now()
    log.append(f"[{fim}] Fim processamento")
    log.append(f"Arquivo salvo em {output_path}")

    # Grava log
    with open(output_path.replace(".xml", "_log_todas_abas.txt"), "w", encoding="utf-8") as f:
        f.write("\n".join(log))

    print("\n".join(log))


if __name__ == "__main__":
    base_dir = r"F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/20 . Carga/Curva C5"  # Ajuste para seu diretÃ³rio

    preencher_todas_abas(
        template_path=os.path.join(base_dir, "Template_Materiais_Criacao_Curva_X.xml"), #Aqui colocar o caminho do template XML
        dados_path=os.path.join(base_dir, "Template_Materiais_Criacao_Curva_C5.xlsx"), # Aqui colocar o caminho do XLSX com os dados
        output_path=os.path.join(base_dir, "Output_Fornecedor_Todas.xml"),# Aqui colocar o caminho do XML de saÃ­da
        row_start_dest=9 # Linha inicial para preenchimento (9 por padrÃ£o, pode ser ajustada
    )