import win32com.client as win32
import openpyxl
import os
import psutil
import time as time_module
import shutil
from datetime import datetime
from tqdm import tqdm  # progress bar

def fechar_processos_excel():
    """Fecha todos os processos do Excel que possam estar rodando"""
    try:
        for proc in psutil.process_iter(['pid', 'name']):
            if proc.info['name'] and 'excel' in proc.info['name'].lower():
                try:
                    proc.terminate()
                    proc.wait(timeout=3)
                    print(f"Processo Excel {proc.info['pid']} terminado")
                except:
                    pass
    except:
        pass

def salvar_arquivo_com_fallback(wb, output_path, log):
    """
    Tenta salvar o arquivo em diferentes formatos até conseguir
    """
    success = False
    
    # Remove arquivo de saída se já existir
    if os.path.exists(output_path):
        try:
            os.remove(output_path)
            log.append(f"🗑️ Arquivo existente removido: {output_path}")
        except Exception as e:
            log.append(f"⚠️ Não foi possível remover arquivo existente: {str(e)}")
    
    # Gera nome único para arquivos temporários
    timestamp = int(time_module.time())
    output_dir = os.path.dirname(output_path)
    temp_xlsx = os.path.join(output_dir, f"temp_xlsx_{timestamp}.xlsx")
    temp_xls = os.path.join(output_dir, f"temp_xls_{timestamp}.xls")
    
    try:
        # Método 1: Salvar diretamente como XML
        log.append("🔄 Tentativa 1: Salvando diretamente como XML...")
        wb.SaveAs(output_path, FileFormat=46)
        log.append(f"✅ Sucesso! Arquivo salvo como XML: {output_path}")
        return True
        
    except Exception as e:
        log.append(f"❌ Falha na tentativa 1: {str(e)}")
        
        try:
            # Método 2: Salvar como XLSX e depois converter
            log.append("🔄 Tentativa 2: Salvando como XLSX temporário...")
            wb.SaveAs(temp_xlsx, FileFormat=51)  # xlsx
            wb.Close(SaveChanges=False)
            
            # Reabre e tenta converter
            excel = wb.Application
            wb_temp = excel.Workbooks.Open(temp_xlsx)
            wb_temp.SaveAs(output_path, FileFormat=46)
            wb_temp.Close(SaveChanges=False)
            
            # Remove arquivo temporário
            if os.path.exists(temp_xlsx):
                os.remove(temp_xlsx)
            
            log.append(f"✅ Sucesso! Arquivo convertido para XML: {output_path}")
            return True
            
        except Exception as e2:
            log.append(f"❌ Falha na tentativa 2: {str(e2)}")
            
            try:
                # Método 3: Salvar como XLS e depois converter
                log.append("🔄 Tentativa 3: Salvando como XLS temporário...")
                if 'wb_temp' in locals():
                    wb_temp.Close(SaveChanges=False)
                
                wb_temp = excel.Workbooks.Open(temp_xlsx) if os.path.exists(temp_xlsx) else wb
                wb_temp.SaveAs(temp_xls, FileFormat=56)  # xls
                wb_temp.Close(SaveChanges=False)
                
                # Reabre XLS e tenta converter
                wb_temp = excel.Workbooks.Open(temp_xls)
                wb_temp.SaveAs(output_path, FileFormat=46)
                wb_temp.Close(SaveChanges=False)
                
                # Remove arquivos temporários
                for temp_file in [temp_xlsx, temp_xls]:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                
                log.append(f"✅ Sucesso! Arquivo convertido para XML via XLS: {output_path}")
                return True
                
            except Exception as e3:
                log.append(f"❌ Falha na tentativa 3: {str(e3)}")
                
                try:
                    # Método 4: Salvar como XLSX final
                    log.append("🔄 Tentativa 4: Salvando como XLSX final...")
                    final_xlsx = output_path.replace('.xml', '.xlsx')
                    
                    if os.path.exists(temp_xlsx):
                        shutil.copy2(temp_xlsx, final_xlsx)
                        log.append(f"⚠️ Arquivo salvo como XLSX: {final_xlsx}")
                        success = True
                    else:
                        # Reabre workbook original se necessário
                        wb.SaveAs(final_xlsx, FileFormat=51)
                        log.append(f"⚠️ Arquivo salvo como XLSX (fallback): {final_xlsx}")
                        success = True
                        
                except Exception as e4:
                    log.append(f"❌ Falha na tentativa 4: {str(e4)}")
                    
                    try:
                        # Método 5: Salvar como XLS final
                        log.append("🔄 Tentativa 5: Salvando como XLS final...")
                        final_xls = output_path.replace('.xml', '.xls')
                        wb.SaveAs(final_xls, FileFormat=56)
                        log.append(f"⚠️ Arquivo salvo como XLS: {final_xls}")
                        success = True
                        
                    except Exception as e5:
                        log.append(f"❌ Falha na tentativa 5: {str(e5)}")
    
    # Limpa arquivos temporários
    for temp_file in [temp_xlsx, temp_xls]:
        if os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
    
    return success

def preencher_todas_abas(template_path, dados_path, output_path, row_start_dest=9):
    """
    Preenche todas as abas do template SAP Migration Cockpit (LTMC/LTMOM) com dados de um XLSX,
    desde que existam abas com o mesmo nome em ambos.
    """

    log = []
    inicio = datetime.now()
    log.append(f"[{inicio}] Início processamento: {os.path.basename(template_path)}")

    # Fecha processos Excel anteriores
    fechar_processos_excel()
    time_module.sleep(1)

    # Carrega XLSX origem
    wb_origem = openpyxl.load_workbook(dados_path, data_only=True)

    # Inicia Excel COM
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False  # Melhora performance
    
    wb_template = excel.Workbooks.Open(template_path)

    abas_xml = [ws.Name for ws in wb_template.Worksheets]
    log.append(f"Abas disponíveis no template XML: {abas_xml}")

    # Loop por todas as abas do XLSX
    for aba_origem in wb_origem.sheetnames:
        ws_origem = wb_origem[aba_origem]

        # Verifica se a aba existe no XML
        if aba_origem not in abas_xml:
            log.append(f"❌ Aba '{aba_origem}' não encontrada no template XML. Pulando...")
            continue

        # Coleta dados a partir da linha definida
        rows = list(ws_origem.iter_rows(min_row=row_start_dest, values_only=True))
        total_rows = len(rows)

        # Verifica se há dados relevantes
        if not any(any(cell is not None for cell in row) for row in rows):
            log.append(f"🔶 Aba '{aba_origem}' ignorada (sem dados na linha {row_start_dest}).")
            continue

        log.append(f"🔵 Preenchendo aba '{aba_origem}' com {total_rows} linhas...")

        ws_template = wb_template.Worksheets(aba_origem)
        
        # Tenta desproteger a planilha
        try:
            ws_template.Unprotect()
        except:
            pass

        row_dest = row_start_dest

        for row in tqdm(rows, desc=f"Preenchendo {aba_origem}", unit="linha"):
            col_dest = 1
            for value in row:
                if value is not None:
                    try:
                        ws_template.Cells(row_dest, col_dest).Value = value
                    except:
                        # Se falhar, tenta como string
                        ws_template.Cells(row_dest, col_dest).Value = str(value)
                col_dest += 1
            row_dest += 1

    # Garante que o diretório de saída existe
    output_dir = os.path.dirname(output_path)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Reativa updates para salvar
    excel.ScreenUpdating = True
    
    # Usa a função de salvamento com fallback
    success = salvar_arquivo_com_fallback(wb_template, output_path, log)
    
    if not success:
        log.append("❌ Falha completa ao salvar arquivo em qualquer formato")

    # Fecha Excel com limpeza forçada
    try:
        # Fecha todos os workbooks abertos
        for wb in excel.Workbooks:
            try:
                wb.Close(SaveChanges=False)
            except:
                pass
        excel.Quit()
        log.append("✅ Excel fechado com sucesso")
    except Exception as e:
        log.append(f"⚠️ Erro ao fechar Excel: {str(e)}")

    fim = datetime.now()
    log.append(f"[{fim}] Fim processamento")
    log.append(f"Tempo total: {fim - inicio}")

    # Grava log
    log_path = output_path.replace(".xml", "_log_todas_abas.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log))

    print("\n".join(log))

if __name__ == "__main__":
    base_dir = r"Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Fornecedor/Conversor"

    preencher_todas_abas(
        template_path=os.path.join(base_dir, "EXP.MIG.LAY.MM-BP01_BP_Fornecedor.xml"),
        dados_path=os.path.join(base_dir, "EXP.MIG.LAY.MM-BP01-BP_For_Preenchido_Par_PF_LP.xlsx"),
        output_path=os.path.join(base_dir, "Output.xml"),
        row_start_dest=9
    )