import openpyxl
import os
import pandas as pd
from sqlalchemy import create_engine
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color

# ====================================================================================
# === CONFIGURAÇÕES GERAIS ==========================================================
# ====================================================================================

# Fonte de dados
usar_sql_server = False  # True para SQL Server, False para arquivos locais

# Conexão SQL Server
sql_server_config = {
    "server": "SEU_SERVIDOR",
    "database": "SEU_BANCO",
    "username": "SEU_USUARIO",
    "password": "SUA_SENHA",
    "driver": "ODBC Driver 17 for SQL Server"
}

# Mapeamento de nomes lógicos para colunas reais
mapeamento_colunas = {
    "PRODUCT": "MATNR",
    "PARTNER": "LIFNR",
    # adicione outros conforme necessário
}


import openpyxl
import os
import pandas as pd
from sqlalchemy import create_engine
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color

# ====================================================================================
# === CONFIGURAÇÕES GERAIS ==========================================================
# ====================================================================================

usar_sql_server = True  # True para SQL Server, False para arquivos locais

sql_server_config = {
    "server": "SPLCPVMSQLQA,23002",
    "database": "MDM_ONDA3",
    "username": "usr_mdm",
    "password": "HSw0raUlpVcBC",
    "driver": "ODBC Driver 17 for SQL Server"
}

# Mapeamento de nomes lógicos para colunas reais
mapeamento_colunas = {
    "PRODUCT": "MATNR",
    "PARTNER": "LIFNR",
    # adicione outros conforme necessário
}


# Caminhos principais
planilha_mapeamento = 'Z:/00 Pastas de trabalho/Asantos/01 - Mapeamento/01 - MM - Materiais/Mapeamento_Material_1.xlsx'
arquivo_destino = 'Z:/00 Pastas de trabalho/Asantos/06 - Template Limpo/Template_limpo_Excel/EXP.MIG.LAY.MM-MM03_Produto.xlsx'
arquivo_destino_preenchido = 'Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Produto/EXP.MIG.LAY.MM-MM03_Produto_Preenchido2.xlsx'
#arquivo_destino_preenchido_LP = 'Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Produto/Template_Materiais_Criacao_Prenchido_LP2.xlsx'
arquivo_log = 'Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Produto/Logs/Log_Processo_Material_Preenchido.xlsx'
Lista_Carga_path = 'Z:/00 Pastas de trabalho/Asantos/08 - Lista_Carga/Lista_Carga_Material_Teste.xlsx'
arquivo_depara = 'f:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/25 . Saneamento/De_Para.xlsx'



# Preparar diretórios
diretorio_destino = os.path.dirname(arquivo_destino_preenchido) or os.getcwd()
if not os.path.exists(diretorio_destino):
    os.makedirs(diretorio_destino)

# Inicializar logs
logs = [["Ação", "Planilha/Tabela", "Coluna de Origem", "Coluna de Destino", "Aba de Destino", "Resultado"]]

# ====================================================================================
# === CARREGAMENTO DE ARQUIVOS E DADOS ==============================================
# ====================================================================================

# Carregar lista de carga como filtro
df_lista_carga = pd.read_excel(Lista_Carga_path, dtype=str)
codigos_lista_carga = set(df_lista_carga.iloc[:,0].str.strip().str.upper())

# Carregar DE-PARA
def carregar_depara(arquivo):
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active
    dict_dp = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        # Extrai apenas as 3 primeiras colunas
        nome_coluna, de, para = row[:3]

        # Ignora linhas vazias ou incompletas
        if nome_coluna is None or de is None or para is None:
            continue

        nome_coluna = str(nome_coluna).strip().upper()
        de = str(de).strip().upper()
        para = str(para).strip()

        if nome_coluna not in dict_dp:
            dict_dp[nome_coluna] = {}
        dict_dp[nome_coluna][de] = para

    return dict_dp

dict_depara = carregar_depara(arquivo_depara)




# Carregar planilha de mapeamento
wb_mapeamento = openpyxl.load_workbook(planilha_mapeamento, data_only=True)
ws_mapeamento = wb_mapeamento["Mapa"] if "Mapa" in wb_mapeamento.sheetnames else wb_mapeamento.active

# Carregar planilha de destino
wb_destino = openpyxl.load_workbook(arquivo_destino)

# ====================================================================================
# === FUNÇÕES AUXILIARES =============================================================
# ====================================================================================

def carregar_tabela_sql_server(nome_tabela, config):
    connection_string = f"mssql+pyodbc://{config['username']}:{config['password']}@{config['server']}/{config['database']}?driver={config['driver']}"
    engine = create_engine(connection_string)
    with engine.connect() as conn:
        df = pd.read_sql_table(nome_tabela, conn)
    return df

def copiar_estilos(origem, destino):
    if origem.font:
        font_color = Color(rgb=origem.font.color.rgb) if origem.font.color and isinstance(origem.font.color.rgb, str) else None
        destino.font = Font(name=origem.font.name, size=origem.font.size, bold=origem.font.bold, italic=origem.font.italic,
                            vertAlign=origem.font.vertAlign, underline=origem.font.underline, strike=origem.font.strike,
                            color=font_color)
    if origem.fill:
        start_color = Color(rgb=origem.fill.start_color.rgb) if origem.fill.start_color and isinstance(origem.fill.start_color.rgb, str) else None
        end_color = Color(rgb=origem.fill.end_color.rgb) if origem.fill.end_color and isinstance(origem.fill.end_color.rgb, str) else None
        destino.fill = PatternFill(fill_type=origem.fill.fill_type, start_color=start_color, end_color=end_color)
    if origem.border:
        destino.border = origem.border
    if origem.alignment:
        destino.alignment = origem.alignment
    if origem.number_format:
        destino.number_format = origem.number_format

# ====================================================================================
# === FUNÇÃO PRINCIPAL DE PREENCHIMENTO ==============================================
# ====================================================================================

def preencher_dados():
    for row in ws_mapeamento.iter_rows(min_row=2, values_only=True):
        nome_aba, planilha_origem, coluna_origem_destino = row
        coluna_origem_destino_upper = coluna_origem_destino.strip().upper()
        coluna_origem_real = mapeamento_colunas.get(coluna_origem_destino_upper, coluna_origem_destino)

        if nome_aba not in wb_destino.sheetnames:
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Aba de destino não encontrada"])
            continue
        ws_destino = wb_destino[nome_aba]

        # Carregar dados origem
        if usar_sql_server:
            df_origem = carregar_tabela_sql_server(planilha_origem, sql_server_config)
            if coluna_origem_real not in df_origem.columns:
                logs.append(["Erro SQL", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna de origem não encontrada"])
                continue
        else:
            caminho_planilha_origem = f"./Extracao_Limpa/{planilha_origem}.xlsx"
            if not os.path.exists(caminho_planilha_origem):
                logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Planilha de origem não encontrada"])
                continue
            wb_origem = openpyxl.load_workbook(caminho_planilha_origem, data_only=False)
            ws_origem = wb_origem.active
            col_origem_idx = next((idx+1 for idx, cell in enumerate(ws_origem[1]) if cell.value and str(cell.value).strip().upper() == coluna_origem_real.strip().upper()), None)
            if col_origem_idx is None:
                logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna de origem não encontrada"])
                continue

        # Localizar coluna destino
        col_dest_idx = next((idx+1 for idx, cell in enumerate(ws_destino[5]) if cell.value and str(cell.value).strip().upper() == coluna_origem_destino_upper), None)
        if col_dest_idx is None:
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna de destino não encontrada"])
            continue

        linha_destino = 9
        if usar_sql_server:
            for _, row_sql in df_origem.iterrows():
                codigo_material = str(row_sql.get("MATNR", "")).strip().upper()
                if codigo_material not in codigos_lista_carga:
                    continue  # ignora se não estiver na lista

                valor = row_sql[coluna_origem_real]
                valor_str = str(valor).strip().upper() if valor else ""
                destino_cell = ws_destino.cell(row=linha_destino, column=col_dest_idx)

                # Aplicar DE-PARA
                de_para_col = dict_depara.get(coluna_origem_real.strip().upper(), {})
                destino_cell.value = de_para_col.get("ALL", de_para_col.get(valor_str, valor))

                linha_destino += 1
        else:
            for row_excel in ws_origem.iter_rows(min_row=2, max_row=ws_origem.max_row):
                valor_origem = row_excel[col_origem_idx - 1]
                if valor_origem:
                    codigo_material = str(row_excel[0].value).strip().upper()
                    if codigo_material not in codigos_lista_carga:
                        continue  # ignora se não estiver na lista

                    destino_cell = ws_destino.cell(row=linha_destino, column=col_dest_idx)
                    valor_str = str(valor_origem.value).strip().upper() if valor_origem.value else ""

                    de_para_col = dict_depara.get(coluna_origem_real.strip().upper(), {})
                    destino_cell.value = de_para_col.get("ALL", de_para_col.get(valor_str, valor_origem.value))

                    copiar_estilos(valor_origem, destino_cell)
                    linha_destino += 1

        logs.append(["Preenchimento concluído", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Sucesso"])

    wb_destino.save(arquivo_destino_preenchido)

# ====================================================================================
# === SALVAR LOGS ===================================================================
# ====================================================================================

def salvar_logs():
    wb_log = openpyxl.Workbook()
    ws_log = wb_log.active
    ws_log.title = "Log"
    for log in logs:
        ws_log.append(log)
    wb_log.save(arquivo_log)

# ====================================================================================
# === EXECUÇÃO PRINCIPAL =============================================================
# ====================================================================================

if __name__ == "__main__":
    preencher_dados()
    salvar_logs()
    print(f"✅ Processamento finalizado. Planilha salva em {arquivo_destino_preenchido}")
