import pandas as pd
from sqlalchemy import create_engine
import urllib
import os
from openpyxl import load_workbook

# === CONFIGURAÇÃO DE CONEXÃO COM SQLALCHEMY ===
conn_str = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=SPLCPVMSQLQA,23002;"
    "DATABASE=MDM_ONDA3;"
    "UID=usr_mdm;"
    "PWD=HSw0raUlpVcBC;"
    "TrustServerCertificate=yes;"
)
conn_str_escaped = urllib.parse.quote_plus(conn_str)
engine = create_engine(f"mssql+pyodbc:///?odbc_connect={conn_str_escaped}")

# === QUERY PARA TRAZER TABELA x COLUNA ===
query = """
select
    tab.name as table_name, 
    col.name as column_name
from sys.tables as tab
    inner join sys.columns as col
        on tab.object_id = col.object_id
"""

# === EXECUÇÃO DA QUERY USANDO SQLALCHEMY ===
df_sql = pd.read_sql(query, engine)

# === CAMINHO DO EXCEL ===
excel_path = r"Z:/00 Pastas de trabalho/Asantos/01 - Mapeamento/Mapeamento_Classes.xlsx"
#excel_path = r"Z:/00 Pastas de trabalho/Asantos/01 - Mapeamento/Mapeamento_Material.xlsx"

if not os.path.exists(excel_path):
    print(f"❌ Arquivo não encontrado: {excel_path}")
    exit(1)
else:
    print(f"✅ Arquivo encontrado: {excel_path}")

# === CARREGAR EXCEL COM openpyxl ===
wb = load_workbook(excel_path)
#ws = wb.active  # ou ws = wb['NomeDaAba'] se precisar aba específica
ws = wb['Mapa']  # acessa a aba real chamada 'Mapa'
# === CARREGAR EM PANDAS PARA FACILIDADE DE TRATAMENTO ===
#df_excel = pd.read_excel(excel_path, engine="openpyxl")
df_excel = pd.read_excel(excel_path, sheet_name="Mapa", engine="openpyxl")
# === MAPEAMENTO: BUSCAR TABELA DE CADA CAMPO ===
def encontrar_tabela(campo):
    resultado = df_sql[df_sql['column_name'].str.upper() == str(campo).strip().upper()]
    if not resultado.empty:
        return ', '.join(resultado['table_name'].unique())
    else:
        return 'NÃO ENCONTRADO'

# === GERAR NOVA COLUNA COM RESULTADO ===
coluna_i_nome = df_excel.columns[2]
nova_coluna = df_excel[coluna_i_nome].apply(encontrar_tabela)

# === ESCREVER NO EXCEL SEM PERDER FORMATO ===
col_destino = ws.max_column + 1
ws.cell(row=1, column=col_destino, value='Tabela_encontrada')  # cabeçalho

for r_idx, value in enumerate(nova_coluna, start=2):  # começa linha 2
    ws.cell(row=r_idx, column=col_destino, value=value)

# === SALVAR ARQUIVO PRESERVANDO LAYOUT ===
output_path = excel_path.replace(".xlsx", "_Tabelas.xlsx")
wb.save(output_path)

print(f"✅ Mapeamento concluído e salvo em: {output_path}")
