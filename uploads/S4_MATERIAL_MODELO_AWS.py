import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color
import psycopg2
import pandas as pd
from sqlalchemy import create_engine

"""
# Caminhos dos arquivos - GITHUB
planilha_mapeamento = './AWS/Mapeamento_Material.xlsx'
arquivo_destino = './AWS/Template_Materiais_Criacao.xlsx'
arquivo_destino_preenchido = './AWS/Template_Materiais_Criacao_DEV.xlsx'
arquivo_log = './AWS/Log_Processo_Material_DEV.xlsx'
Lista_Carga = './AWS/Lista_Material.xlsx'
"""

# Caminhos dos arquivos - PC HOME
planilha_mapeamento = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Mapeamento_Material.xlsx'
arquivo_destino = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Template_Materiais_Criacao.xlsx'
arquivo_destino_preenchido = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Template_Materiais_Criacao_DEV.xlsx'
arquivo_destino_preenchido_LP = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Template_Materiais_Criacao_DEV_LP.xlsx'
arquivo_log = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Log_Processo_Material_DEV.xlsx'
Lista_Carga = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Lista_Material.xlsx'

# Garantir que o diretório de destino exista
for caminho in [arquivo_destino_preenchido, arquivo_log]:
    diretorio = os.path.dirname(caminho) or os.getcwd()
    if not os.path.exists(diretorio):
        print(f"Diretório '{diretorio}' não encontrado. Criando...")
        os.makedirs(diretorio)

# Lista de logs
logs = [["Ação", "Planilha de Origem", "Coluna de Origem", "Coluna de Destino", "Aba de Destino", "Resultado"]]

# Abrir planilha de mapeamento e selecionar aba "Mapa"
wb_mapeamento = openpyxl.load_workbook(planilha_mapeamento, data_only=True)
ws_mapeamento = wb_mapeamento["Mapa"] if "Mapa" in wb_mapeamento.sheetnames else None
if ws_mapeamento is None:
    raise ValueError("A aba 'Mapa' não foi encontrada na planilha de mapeamento.")

# Abrir a planilha de destino
wb_destino = openpyxl.load_workbook(arquivo_destino)

def reexibir_linha_5(wb):
    for aba in wb.sheetnames:
        if aba not in ["Introdução", "Lista de campos"]:
            ws = wb[aba]
            if ws.row_dimensions[5].hidden:
                ws.row_dimensions[5].hidden = False
                print(f"Linha 5 reexibida na aba '{aba}'.")
                logs.append(["Linha reexibida", "-", "-", "-", aba, "Sucesso"])

def copiar_estilos(origem, destino):
    if origem.font:
        destino.font = Font(**{k: getattr(origem.font, k) for k in vars(origem.font) if hasattr(origem.font, k)})
    if origem.fill:
        destino.fill = PatternFill(fill_type=origem.fill.fill_type,
                                   start_color=origem.fill.start_color,
                                   end_color=origem.fill.end_color)
    if origem.border:
        destino.border = Border(**{k: getattr(origem.border, k) for k in vars(origem.border) if hasattr(origem.border, k)})
    if origem.alignment:
        destino.alignment = Alignment(**{k: getattr(origem.alignment, k) for k in vars(origem.alignment) if hasattr(origem.alignment, k)})
    destino.number_format = origem.number_format

# Carregar a planilha Lista_Carga (caso exista) e ler os códigos de MATNR com 18 dígitos
def carregar_lista_carga():
    if not os.path.exists(Lista_Carga):
        print(f"⚠️ Arquivo '{Lista_Carga}' não encontrado.")
        return set()

    try:
        df_lista = pd.read_excel(Lista_Carga, sheet_name=0)  # Primeira aba
        if "MATNR" not in df_lista.columns:
            print(f"⚠️ Coluna 'MATNR' não encontrada em {Lista_Carga}.")
            return set()

        codigos = (
            df_lista["MATNR"]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
            .str.zfill(18)  # Preenche com zeros à esquerda até 18 dígitos
            .tolist()
        )
        return set(codigos)

    except Exception as e:
        print(f"❌ Erro ao carregar '{Lista_Carga}': {e}")
        return set()

# Uso:
codigos = carregar_lista_carga()


from sqlalchemy import create_engine

def preencher_dados():
    print("Códigos carregados:", codigos)
    print("Total de códigos:", len(codigos))
    lista_carga = carregar_lista_carga()
    engine = create_engine("postgresql+psycopg2://postgres:stra17xx%$1962@analista26.c3csoio6kacg.us-east-2.rds.amazonaws.com:5432/analista26")


    from sqlalchemy import text

    for row in ws_mapeamento.iter_rows(min_row=2, values_only=True):
        nome_aba, planilha_origem, coluna_origem_destino = row
        coluna_origem_destino_upper = coluna_origem_destino.strip().upper()

        # Mapeamento correto para "MATNR"
        if coluna_origem_destino_upper in ["PRODUCT", "BISMT"]:
            coluna_origem_real = "MATNR"
        else:
            coluna_origem_real = coluna_origem_destino

        with engine.connect() as conn:
            try:
                # Verifica se a tabela existe no banco
                query = text("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name=:table_name)")
                result = conn.execute(query, {"table_name": planilha_origem})
                if not result.fetchone()[0]:
                    print(f"⚠️ Tabela '{planilha_origem}' não encontrada no banco. Pulando...")
                    logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Tabela não encontrada no banco"])
                    continue

                # Lê a tabela inteira (você pode limitar colunas se quiser)
                df_origem = pd.read_sql_query(f'SELECT * FROM "{planilha_origem}"', conn)

            except Exception as e:
                print(f"❌ Erro ao ler tabela '{planilha_origem}': {e}")
                logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, f"Erro ao ler tabela: {e}"])
                continue 
        
        
    
        if lista_carga and "MATNR" in df_origem.columns:
            df_origem = df_origem[df_origem["MATNR"].astype(str).str.upper().isin(lista_carga)]

        if coluna_origem_real not in df_origem.columns:
            print(f"⚠️ Coluna '{coluna_origem_real}' não encontrada na tabela '{planilha_origem}'. Pulando...")
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna não encontrada"])
            continue

        if nome_aba not in wb_destino.sheetnames:
            print(f"⚠️ Aba '{nome_aba}' não encontrada na planilha de destino. Pulando...")
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Aba de destino não encontrada"])
            continue

        ws_destino = wb_destino[nome_aba]
        col_dest_idx = None
        for col_idx, cell in enumerate(ws_destino[5], start=1):
            if cell.value and str(cell.value).strip().upper() == coluna_origem_destino.strip().upper():
                col_dest_idx = col_idx
                break

        if col_dest_idx is None:
            print(f"⚠️ Coluna '{coluna_origem_destino}' não encontrada na aba '{nome_aba}'. Pulando...")
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna de destino não encontrada"])
            continue

        print(f"✅ Preenchendo dados da coluna '{coluna_origem_real}' da tabela '{planilha_origem}' para a aba '{nome_aba}'.")
        logs.append(["Preenchimento iniciado", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Sucesso"])

        substituicoes = {"ZSPR": "ERSA", "ZFER": "FERT", "ZOPR": "FHMI", "ZHAL": "HALB", "ZREV": "HAWA",
                         "ZCON": "HIBE", "ZHIB": "HIBE", "ZEMB": "LEIH", "ZLAG": "NLAG", "ZRAW": "ROH",
                         "ZROH": "ROH", "ZDIE": "SERV", "ZSER": "SERV", "ZATI": "UNBW", "ZEMB": "VERP",
                         "ZINT": "ZINT"}

        Centro = {"6001": "9501", "6002": "9502", "6003": "9503", "6004": "9504", "6005": "9505",
                  "6006": "9506", "6007": "9507", "6008": "9508", "6010": "9510", "6012": "9511",
                  "6013": "9512", "6014": "9513", "6015": "9514", "6016": "9515", "6017": "9516",
                  "6018": "9517", "6019": "9518", "6020": "9519", "6021": "9520", "6022": "9521",
                  "6023": "9522", "6024": "9523", "6101": "9524", "6202": "9525", "6203": "9526",
                  "6204": "9527"}

        linha_destino = 9
        for _, row_data in df_origem.iterrows():
            valor = str(row_data[coluna_origem_real]).strip().upper() if pd.notna(row_data[coluna_origem_real]) else ""
            if valor:
                cell = ws_destino.cell(row=linha_destino, column=col_dest_idx)
                if coluna_origem_real.upper() == "MTART":
                    cell.value = substituicoes.get(valor, valor)
                elif coluna_origem_real.upper() in ["WERKS", "BWKEY"]:
                    cell.value = Centro.get(valor, valor)
                else:
                    cell.value = valor
                linha_destino += 1

        col_waers_idx = col_spras_idx = col_product_idx = None
        for idx, cell in enumerate(ws_destino[5]):
            nome = str(cell.value).strip().upper() if cell.value else ""
            if nome == "WAERS": col_waers_idx = idx + 1
            elif nome == "SPRAS": col_spras_idx = idx + 1
            elif nome == "PRODUCT": col_product_idx = idx + 1

        for linha in range(9, ws_destino.max_row + 1):
            if col_product_idx and ws_destino.cell(row=linha, column=col_product_idx).value:
                if col_waers_idx: ws_destino.cell(row=linha, column=col_waers_idx).value = "BRL"
                if col_spras_idx: ws_destino.cell(row=linha, column=col_spras_idx).value = "PT"

    wb_destino.save(arquivo_destino_preenchido)
    print(f"Planilha preenchida e salva em: {arquivo_destino_preenchido}")

def salvar_logs():
    wb_log = openpyxl.Workbook()
    ws_log = wb_log.active
    ws_log.title = "Log"
    for log in logs:
        ws_log.append(log)
    wb_log.save(arquivo_log)
    print(f"Logs salvos em: {arquivo_log}")

reexibir_linha_5(wb_destino)
preencher_dados()
salvar_logs()    

###############################################**** FORMATO ****#################################################

from openpyxl import load_workbook
from openpyxl.styles import numbers
import re
from tqdm import tqdm  # <-- Importa a barra de progresso

# Carrega a planilha
wb = load_workbook(arquivo_destino_preenchido)

# Regex para identificar valores numéricos com vírgulas ou pontos
regex_numerico = re.compile(r"^[\d\.,]+$")

def converter_valor(celula):
    valor_original = str(celula.value).strip()
    
    if celula.value is None or isinstance(celula.value, (int, float)):
        return
    if valor_original.isdigit():
        return
    if regex_numerico.match(valor_original):
        valor_convertido = valor_original.replace('.', '').replace(',', '.')
        try:
            numero = float(valor_convertido)
            celula.value = numero
            casas_decimais = len(valor_original.split(',')[1]) if ',' in valor_original else 0
            formato_excel = f'#,##0{"." + "0"*casas_decimais if casas_decimais else ""}'
            celula.number_format = formato_excel
        except ValueError:
            pass

# Calcular total de células para mostrar progresso
total_celulas = sum(len(linha) for aba in wb.worksheets for linha in aba.iter_rows())

# Processar com tqdm (barra de progresso)
with tqdm(total=total_celulas, desc="Processando células", ncols=100) as pbar:
    for aba in wb.worksheets:
        for linha in aba.iter_rows():
            for celula in linha:
                converter_valor(celula)
                pbar.update(1)

# Salvar o novo arquivo
wb.save(arquivo_destino_preenchido_LP)
print(f"\n✅ Arquivo salvo como: {arquivo_destino_preenchido_LP}")

#############################################*** LIMPEZA ***###############################################

import openpyxl
from tqdm import tqdm


# Carregar o arquivo Excel
wb = openpyxl.load_workbook(arquivo_destino_preenchido_LP)

# Lista expandida de formatos que representam zero decimal
formatos_zero = [
    "0,000", "0,00", "0,0", "#,##0.000", "#,##0.00", "#,##0.0",  
    "#,##0", "#,###0", "#.##0,00", "#.##0,0", "#.##0", 
    "0.000", "0.00", "0.0"
]

# Processar todas as abas
for aba in tqdm(wb.sheetnames, desc="Processando abas"):
    ws = wb[aba]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value == 0:
                # Capturar o formato da célula e remover aspas se existirem
                formato = cell.number_format.replace('"', '').strip()
                
                # Se o formato da célula estiver na lista de zeros formatados, remover valor
                if any(f in formato for f in formatos_zero):
                    cell.value = None  # Apaga o valor sem alterar o formato

# Salvar mantendo o layout original
wb.save(arquivo_destino_preenchido_LP)

print("✅ Processamento concluído. Arquivo salvo em:",arquivo_destino_preenchido_LP )