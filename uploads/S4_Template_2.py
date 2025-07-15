import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color
import psycopg2
import pandas as pd
from sqlalchemy import create_engine


# Caminhos dos arquivos - GITHUB
#planilha_mapeamento = './AWS/Mapeamento_Material.xlsx'
#arquivo_destino = './AWS/Template_Materiais_Criacao.xlsx'
#arquivo_destino_preenchido = './AWS/Template_Materiais_Criacao_DEV.xlsx'
#arquivo_destino_preenchido_LP = './AWS/Template_Materiais_Criacao_DEV.xlsx'
#arquivo_log = './AWS/Log_Processo_Material_DEV.xlsx'
#Lista_Carga = './AWS/Lista_Material.xlsx'

# Caminhos dos arquivos - PC HOME
#planilha_mapeamento = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Mapeamento_Material.xlsx'
#arquivo_destino = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Template_Materiais_Criacao.xlsx'
#arquivo_destino_preenchido = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Template_Materiais_Criacao_DEV.xlsx'
#arquivo_destino_preenchido_LP = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Template_Materiais_Criacao_DEV_LP.xlsx'
#arquivo_log = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Log_Processo_Material_DEV.xlsx'
#Lista_Carga = 'C:/Users/HOME/Desktop/ONDA_02/AWS/Lista_Material.xlsx'

# Caminhos dos arquivos - PC EMPRESA
#Caminhos dos arquivos - PC EMPRESA - Fornecedor - # XXX AJUSTE AQUI XXX
planilha_mapeamento = 'F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/28 . Automatizacao/Fornecedor/Planilha_Mapeamento_Fornecedor_2.xlsx'  # OK
arquivo_destino = 'F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/28 . Automatizacao/Template Limpo/Template_limpo_Excel/EXP.MIG.LAY.MM-BP01-BP_Fornecedor.xlsx' # OK
arquivo_destino_preenchido = 'F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/28 . Automatizacao/Fornecedor/05 - X_TEMPLATE_PREENCHIDO/BP_Fornecedor_Preenchido2.xlsx'  # OK
arquivo_destino_preenchido_LP = 'F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/28 . Automatizacao/Fornecedor//05 - X_TEMPLATE_PREENCHIDO/BP_Fornecedor_Preenchido_LP2.xlsx' # OK
arquivo_log = 'F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/28 . Automatizacao/Fornecedor//05 - X_TEMPLATE_PREENCHIDO/Logs/Log_Processo_Fornecedor_Preenchido.xlsx' # OK
Lista_Carga = 'F:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/28 . Automatizacao/Fornecedor/Lista_Carga_Fornecedor.xlsx' #OK

"""
planilha_mapeamento = r"E:/Dev/Mapeamento_Material.xlsx"
arquivo_destino = r"E:/Dev/EXP.MIG.LAY.MM-MM03 - Produto.xlsx"
arquivo_destino_preenchido = r"E:/Dev/EXP.MIG.LAY.MM-MM03 - Produto_preenchido.xlsx"
arquivo_destino_preenchido_LP = r"E:/Dev/EXP.MIG.LAY.MM-MM03 - Produto_preenchido_LP.xlsx"
arquivo_log = r"E:/Dev/Log_Processo_Material_DEV.xlsx"
Lista_Carga = r"E:/Dev/Lista_Material.xlsx"
"""


# Garantir que o diretório de destino exista
for caminho in [arquivo_destino_preenchido, arquivo_log]:
    diretorio = os.path.dirname(caminho) or os.getcwd()
    if not os.path.exists(diretorio):
        print(f"Diretório '{diretorio}' não encontrado. Criando...")
        os.makedirs(diretorio)

# Lista de logs
logs = [["Ação", "Planilha de Origem", "Coluna de Origem", "Coluna de Destino", "Aba de Destino", "Resultado"]]

# Abrir planilha de mapeamento e selecionar aba "Mapa"
wb_mapeamento = openpyxl.load_workbook(planilha_mapeamento, data_only=True)
ws_mapeamento = wb_mapeamento["Mapa"] if "Mapa" in wb_mapeamento.sheetnames else None
if ws_mapeamento is None:
    raise ValueError("A aba 'Mapa' não foi encontrada na planilha de mapeamento.")

# Abrir a planilha de destino
wb_destino = openpyxl.load_workbook(arquivo_destino)

def reexibir_linha_5(wb):
    for aba in wb.sheetnames:
        if aba not in ["Introdução", "Lista de campos"]:
            ws = wb[aba]
            if ws.row_dimensions[5].hidden:
                ws.row_dimensions[5].hidden = False
                print(f"Linha 5 reexibida na aba '{aba}'.")
                logs.append(["Linha reexibida", "-", "-", "-", aba, "Sucesso"])

def copiar_estilos(origem, destino):
    if origem.font:
        destino.font = Font(**{k: getattr(origem.font, k) for k in vars(origem.font) if hasattr(origem.font, k)})
    if origem.fill:
        destino.fill = PatternFill(fill_type=origem.fill.fill_type,
                                   start_color=origem.fill.start_color,
                                   end_color=origem.fill.end_color)
    if origem.border:
        destino.border = Border(**{k: getattr(origem.border, k) for k in vars(origem.border) if hasattr(origem.border, k)})
    if origem.alignment:
        destino.alignment = Alignment(**{k: getattr(origem.alignment, k) for k in vars(origem.alignment) if hasattr(origem.alignment, k)})
    destino.number_format = origem.number_format

############################################################## LISTA P/CARGA #####################################################

# Carregar a planilha Lista_Carga (caso exista) e ler os códigos de MATNR com 18 dígitos
def carregar_lista_carga():
    if not os.path.exists(Lista_Carga):
        print(f"[ERRO] Arquivo '{Lista_Carga}' não encontrado.")
        return set()

    try:
        df_lista = pd.read_excel(Lista_Carga, sheet_name=0)  # Primeira aba
        if "MATNR" not in df_lista.columns:
            print(f"[WARN] Coluna 'MATNR' não encontrada em {Lista_Carga}.")
            return set()

        codigos = (
            df_lista["MATNR"]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
            .str.zfill(18)  # Preenche com zeros à esquerda até 18 dígitos
            .tolist()
        )
        return set(codigos)

    except Exception as e:
        print(f"[ERRO] Erro ao carregar '{Lista_Carga}': {e}")
        return set()

# Uso:
codigos = carregar_lista_carga()


###############################################**** Preencher_Dados_nova ****#####################################################

from sqlalchemy import create_engine, text
import urllib
import pandas as pd

def buscar_valor_mapeado(coluna, valor_original, df_de_para):
    valor_original = str(valor_original).strip().upper()
    mapeado = df_de_para[
        (df_de_para["COLUNA"].str.upper() == coluna.upper()) & 
        (df_de_para["DE"].astype(str).str.upper() == valor_original)
    ]
    if not mapeado.empty:
        return mapeado.iloc[0]["PARA"]
    mapeado_all = df_de_para[
        (df_de_para["COLUNA"].str.upper() == coluna.upper()) & 
        (df_de_para["DE"].astype(str).str.upper() == "ALL")
    ]
    if not mapeado_all.empty:
        return mapeado_all.iloc[0]["PARA"]
    return valor_original

def preencher_dados():
    print("Códigos carregados:", codigos)
    print("Total de códigos:", len(codigos))
    lista_carga = carregar_lista_carga()

    conn_str = (
        "DRIVER={ODBC Driver 17 for SQL Server};"
        "SERVER=SPLCPVMSQLQA,23002;"
        "DATABASE=MDM_ONDA3;"
        "UID=usr_mdm;"
        "PWD=HSw0raUlpVcBC;"
        "TrustServerCertificate=yes;"
    )
    conn_str_escaped = urllib.parse.quote_plus(conn_str)
    engine = create_engine(f"mssql+pyodbc:///?odbc_connect={conn_str_escaped}")  


    df_de_para = pd.read_excel(r"f:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/25 . Saneamento/De_Para.xlsx")  # Substitua pelo caminho real

    for row in ws_mapeamento.iter_rows(min_row=2, values_only=True):
        nome_aba, planilha_origem, coluna_origem_destino = row
        coluna_origem_destino_upper = coluna_origem_destino.strip().upper()
        coluna_origem_real = "MATNR" if coluna_origem_destino_upper in ["PRODUCT", "BISMT"] else coluna_origem_destino

        with engine.connect() as conn:
            try:
                query = text("""
                    SELECT CASE 
                        WHEN EXISTS (
                            SELECT * FROM INFORMATION_SCHEMA.TABLES 
                            WHERE TABLE_NAME = :table_name AND TABLE_SCHEMA = 'dbo'
                        ) THEN CAST(1 AS BIT)
                        ELSE CAST(0 AS BIT) 
                    END
                """)
                result = conn.execute(query, {"table_name": planilha_origem})
                if not result.fetchone()[0]:
                    print(f"[WARN] Tabela '{planilha_origem}' não encontrada no banco. Pulando...")
                    logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Tabela não encontrada no banco"])
                    continue
                df_origem = pd.read_sql_query(f'SELECT * FROM "{planilha_origem}"', conn)
            except Exception as e:
                print(f"[ERRO] Erro ao ler tabela '{planilha_origem}': {e}")
                logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, f"Erro ao ler tabela: {e}"])
                continue

        if lista_carga and "MATNR" in df_origem.columns:
            df_origem = df_origem[df_origem["MATNR"].astype(str).str.upper().isin(lista_carga)]

        if coluna_origem_real not in df_origem.columns:
            print(f"[WARN] Coluna '{coluna_origem_real}' não encontrada na tabela '{planilha_origem}'. Pulando...")
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna não encontrada"])
            continue

        if nome_aba not in wb_destino.sheetnames:
            print(f"[WARN] Aba '{nome_aba}' não encontrada na planilha de destino. Pulando...")
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Aba de destino não encontrada"])
            continue

        ws_destino = wb_destino[nome_aba]
        col_dest_idx = None
        for col_idx, cell in enumerate(ws_destino[5], start=1):
            if cell.value and str(cell.value).strip().upper() == coluna_origem_destino_upper:
                col_dest_idx = col_idx
                break

        if col_dest_idx is None:
            print(f"[WARN] Coluna '{coluna_origem_destino}' não encontrada na aba '{nome_aba}'. Pulando...")
            logs.append(["Erro", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Coluna de destino não encontrada"])
            continue

        print(f"[OK]Preenchendo dados da coluna '{coluna_origem_real}' da tabela '{planilha_origem}' para a aba '{nome_aba}'.")
        logs.append(["Preenchimento iniciado", planilha_origem, coluna_origem_real, coluna_origem_destino, nome_aba, "Sucesso"])

        linha_destino = 9
        for _, row_data in df_origem.iterrows():
            valor_original = row_data.get(coluna_origem_real)
            if pd.notna(valor_original):
                valor_formatado = str(valor_original).strip().upper()
                valor_mapeado = buscar_valor_mapeado(coluna_origem_real, valor_formatado, df_de_para)
                ws_destino.cell(row=linha_destino, column=col_dest_idx).value = valor_mapeado
                linha_destino += 1

    wb_destino.save(arquivo_destino_preenchido)
    print(f"Planilha preenchida e salva em: {arquivo_destino_preenchido}")

def salvar_logs():
    wb_log = openpyxl.Workbook()
    ws_log = wb_log.active
    ws_log.title = "Log"
    for log in logs:
        ws_log.append(log)
    wb_log.save(arquivo_log)
    print(f"Logs salvos em: {arquivo_log}")

reexibir_linha_5(wb_destino)
preencher_dados()
salvar_logs()    





###############################################**** FORMATO ****###############################################################
from openpyxl import load_workbook
from openpyxl.styles import numbers
import re
from tqdm import tqdm  # <-- Importa a barra de progresso

# Carrega a planilha
wb = load_workbook(arquivo_destino_preenchido)

# Regex para identificar valores numéricos com vírgulas ou pontos
regex_numerico = re.compile(r"^[\d\.,]+$")

def converter_valor(celula):
    valor_original = str(celula.value).strip()
    
    if celula.value is None or isinstance(celula.value, (int, float)):
        return
    if valor_original.isdigit():
        return
    if regex_numerico.match(valor_original):
        valor_convertido = valor_original.replace('.', '').replace(',', '.')
        try:
            numero = float(valor_convertido)
            celula.value = numero
            casas_decimais = len(valor_original.split(',')[1]) if ',' in valor_original else 0
            formato_excel = f'#,##0{"." + "0"*casas_decimais if casas_decimais else ""}'
            celula.number_format = formato_excel
        except ValueError:
            pass

# Calcular total de células para mostrar progresso
total_celulas = sum(len(linha) for aba in wb.worksheets for linha in aba.iter_rows())

# Processar com tqdm (barra de progresso)
with tqdm(total=total_celulas, desc="Processando células", ncols=100) as pbar:
    for aba in wb.worksheets:
        for linha in aba.iter_rows():
            for celula in linha:
                converter_valor(celula)
                pbar.update(1)

# Salvar o novo arquivo
wb.save(arquivo_destino_preenchido_LP)
print(f"\n[OK] Arquivo salvo como: {arquivo_destino_preenchido_LP}")

#############################################*** LIMPEZA ***###############################################

import openpyxl
from tqdm import tqdm


# Carregar o arquivo Excel
wb = openpyxl.load_workbook(arquivo_destino_preenchido_LP)

# Lista expandida de formatos que representam zero decimal
formatos_zero = [
    "0,000", "0,00", "0,0", "#,##0.000", "#,##0.00", "#,##0.0",  
    "#,##0", "#,###0", "#.##0,00", "#.##0,0", "#.##0", 
    "0.000", "0.00", "0.0"
]

# Processar todas as abas
for aba in tqdm(wb.sheetnames, desc="Processando abas"):
    ws = wb[aba]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value == 0:
                # Capturar o formato da célula e remover aspas se existirem
                formato = cell.number_format.replace('"', '').strip()
                
                # Se o formato da célula estiver na lista de zeros formatados, remover valor
                if any(f in formato for f in formatos_zero):
                    cell.value = None  # Apaga o valor sem alterar o formato

# Salvar mantendo o layout original
wb.save(arquivo_destino_preenchido_LP)

print("[OK] Processamento concluído. Arquivo salvo em:",arquivo_destino_preenchido_LP )
