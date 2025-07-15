import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment, Color
import psycopg2
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy import create_engine, text
import urllib
import pandas as pd

############################################################################################################################
ogs = []  # inicializa lista de logs

# ================== CONFIGURAÇÃO DE MODO DE EXECUÇÃO ==================
usar_sql_server = True  # True para usar SQL Server; False para ler planilhas locais
#usar_sql_server = False  # True para usar SQL Server; False para ler planilhas locais

# ================================
# CONFIGURAÇÃO DE EXECUÇÃO DINÂMICA
# ================================

# Define o tipo de execução
tipo_execucao = "produto"  # opções: "produto", "bp_fornecedor", "bp_cliente"
print(f"Execução configurada para: {tipo_execucao.upper()}")

# Define coluna_chave de acordo com tipo_execucao (Produto, Cliente, Fornecedor)
coluna_chave_execucao = {
    "produto": "MATNR",
    "bp_fornecedor": "LIFNR",
    "bp_cliente": "KUNNR"
}

coluna_chave = coluna_chave_execucao[tipo_execucao]
print(f"Coluna-chave definida como: {coluna_chave}")

#=========================================== CAMINHOS ===========================================

# Caminhos de arquivos por tipo_execucao
caminhos_config = {
    "produto": {
        "planilha_mapeamento": 'Z:/00 Pastas de trabalho/Asantos/01 - Mapeamento/01 - MM - Materiais/Mapeamento_Material.xlsx',
        "arquivo_destino": 'Z:/00 Pastas de trabalho/Asantos/06 - Template Limpo/Template_limpo_Excel/EXP.MIG.LAY.MM-MM03_Produto.xlsx',
        "arquivo_destino_preenchido": 'Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Produto/EXP.MIG.LAY.MM-MM03_Produto_Preenchido.xlsx',
        "arquivo_destino_preenchido_LP": 'Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Produto/Template_Materiais_Criacao_Prenchido_LP.xlsx',
        "arquivo_log": 'Z:/00 Pastas de trabalho/Asantos/05 - X_TEMPLATE_PREENCHIDO/Produto/Logs/Log_Processo_Material_Preenchido.xlsx',
        "Lista_Carga": 'Z:/00 Pastas de trabalho/Asantos/08 - Lista_Carga/Lista_Carga_Material_Teste.xlsx',
    },
    "bp_fornecedor": {
        "planilha_mapeamento": 'E:/Dev/Fornecedor/Planilha_Mapeamento_Fornecedor_2.xlsx',
        "arquivo_destino": 'E:/Dev/Fornecedor/BP_Fornecedor.xlsx',
        "arquivo_destino_preenchido": 'E:/Dev/Fornecedor/X_TEMPLATE_PREENCHIDO/BP_Fornecedor_Preenchido.xlsx',
        "arquivo_destino_preenchido_LP": 'E:/Dev/Fornecedor/X_TEMPLATE_PREENCHIDO/BP_Fornecedor_Preenchido_LP.xlsx',
        "arquivo_log": 'E:/Dev/Fornecedor/X_TEMPLATE_PREENCHIDO/Logs/Log_Processo_Fornecedor_Preenchido.xlsx',
        "Lista_Carga": 'E:/Dev/Fornecedor/Lista_Carga_Fornecedor.xlsx',
    },
    "bp_cliente": {
        "planilha_mapeamento": 'CAMINHO/Mapeamento_Cliente.xlsx',
        "arquivo_destino": 'CAMINHO/BP_Cliente.xlsx',
        "arquivo_destino_preenchido": 'CAMINHO/BP_Cliente_Preenchido.xlsx',
        "arquivo_destino_preenchido_LP": 'CAMINHO/BP_Cliente_Preenchido_LP.xlsx',
        "arquivo_log": 'CAMINHO/Log_Processo_Cliente_Preenchido.xlsx',
        "Lista_Carga": 'CAMINHO/Lista_Carga_Cliente.xlsx',
    },
}

# Carregar caminhos conforme tipo_execucao
cfg = caminhos_config[tipo_execucao]
planilha_mapeamento = cfg["planilha_mapeamento"]
arquivo_destino = cfg["arquivo_destino"]
arquivo_destino_preenchido = cfg["arquivo_destino_preenchido"]
arquivo_destino_preenchido_LP = cfg["arquivo_destino_preenchido_LP"]
arquivo_log = cfg["arquivo_log"]
Lista_Carga = cfg["Lista_Carga"]


##############################################################################################################################################################################
def resolver_coluna_origem(coluna_destino, df_origem, mapeamento_excecoes, colunas_equivalentes=None):
    """
    Resolve a coluna de origem real considerando:
    - Mapeamentos de exceções diretas
    - Colunas equivalentes dinâmicas (ex: PRODUCT ↔ MATNR)
    - Fallback para o próprio nome da coluna destino
    """
    coluna_destino_upper = coluna_destino.strip().upper()

    # 1. Verifica mapeamento de exceções diretas
    if coluna_destino_upper in mapeamento_excecoes:
        coluna_mapeada = mapeamento_excecoes[coluna_destino_upper]
        if coluna_mapeada in df_origem.columns:
            return coluna_mapeada

    # 2. Verifica colunas equivalentes (dinâmico)
    if colunas_equivalentes and coluna_destino_upper in colunas_equivalentes:
        for col_equiv in colunas_equivalentes[coluna_destino_upper]:
            if col_equiv in df_origem.columns:
                return col_equiv

    # 3. Se nada encontrado, retorna o próprio nome da coluna destino
    if coluna_destino_upper in df_origem.columns:
        return coluna_destino_upper

    # 4. Se ainda não encontrado, retorna None para tratamento posterior
    return None

###########################################################################################################################################################################
def eh_codigo_com_zeros(valor):
    """
    Detecta se é um código que deve preservar zeros à esquerda
    Critérios mais flexíveis:
    1. String só com dígitos
    2. Começa com zero OU tem padrão de código (ex: 0001, 000, etc.)
    """
    if valor is None:
        return False
    
    valor_str = str(valor).strip()
    
    # Deve ser só dígitos
    if not valor_str.isdigit():
        return False
    
    # Se começa com zero e tem mais de 1 dígito, preservar
    if valor_str.startswith('0') and len(valor_str) > 1:
        return True
    
    # Se tem tamanho fixo comum de códigos (4, 6, 8, 10, 18 dígitos) E começa com zero
    if len(valor_str) in [4, 6, 8, 10, 18] and valor_str.startswith('0'):
        return True
    
    return False

def preservar_formato_original(valor):
    """
    Preserva o formato original se for código com zeros
    """
    if eh_codigo_com_zeros(valor):
        return str(valor)
    return valor

##########################################################################################################################################################################
# Garantir que o diretório de destino exista
for caminho in [arquivo_destino_preenchido, arquivo_log]:
    diretorio = os.path.dirname(caminho) or os.getcwd()
    if not os.path.exists(diretorio):
        print(f"Diretório '{diretorio}' não encontrado. Criando...")
        os.makedirs(diretorio)

# Lista de logs
logs = [["Ação", "Planilha de Origem", "Coluna de Origem", "Coluna de Destino", "Aba de Destino", "Resultado"]]

# Abrir planilha de mapeamento e selecionar aba "Mapa"
wb_mapeamento = openpyxl.load_workbook(planilha_mapeamento, data_only=True)
ws_mapeamento = wb_mapeamento["Mapa"] if "Mapa" in wb_mapeamento.sheetnames else None
if ws_mapeamento is None:
    raise ValueError("A aba 'Mapa' não foi encontrada na planilha de mapeamento.")

# Abrir a planilha de destino
wb_destino = openpyxl.load_workbook(arquivo_destino)

def reexibir_linha_5(wb):
    for aba in wb.sheetnames:
        if aba not in ["Introdução", "Lista de campos"]:
            ws = wb[aba]
            if ws.row_dimensions[5].hidden:
                ws.row_dimensions[5].hidden = False
                print(f"Linha 5 reexibida na aba '{aba}'.")
                logs.append(["Linha reexibida", "-", "-", "-", aba, "Sucesso"])

def copiar_estilos(origem, destino):
    if origem.font:
        destino.font = Font(**{k: getattr(origem.font, k) for k in vars(origem.font) if hasattr(origem.font, k)})
    if origem.fill:
        destino.fill = PatternFill(fill_type=origem.fill.fill_type,
                                   start_color=origem.fill.start_color,
                                   end_color=origem.fill.end_color)
    if origem.border:
        destino.border = Border(**{k: getattr(origem.border, k) for k in vars(origem.border) if hasattr(origem.border, k)})
    if origem.alignment:
        destino.alignment = Alignment(**{k: getattr(origem.alignment, k) for k in vars(origem.alignment) if hasattr(origem.alignment, k)})
    destino.number_format = origem.number_format


def deve_preservar_zeros(valor):
    """
    Detecta dinamicamente se um valor deve preservar zeros à esquerda
    Critérios:
    1. É string composta apenas por dígitos
    2. Começa com zero(s)
    3. Tem mais de 1 dígito
    """
    if not isinstance(valor, str):
        return False
    
    valor_clean = str(valor).strip()
    
    # Deve ser só dígitos
    if not valor_clean.isdigit():
        return False
    
    # Deve começar com zero e ter mais de 1 dígito
    if valor_clean.startswith('0') and len(valor_clean) > 1:
        return True
    
    return False
############################################################## LISTA P/CARGA #####################################################

# Carregar a planilha Lista_Carga (caso exista) e ler os códigos de MATNR com 18 dígitos

def carregar_lista_carga(coluna_chave="MATNR"): # # XXX AJUSTE AQUI XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX XXX 
    """
    Carrega a Lista_Carga considerando a coluna_chave (MATNR, KUNNR, LIFNR) e
    retorna um set com os códigos formatados.

    :param coluna_chave: Nome da coluna-chave no Excel (ex: MATNR, KUNNR, LIFNR)
    :return: set de códigos formatados
    """
    if not os.path.exists(Lista_Carga):
        print(f"Arquivo '{Lista_Carga}' não encontrado.")
        return set()

    try:
        df_lista = pd.read_excel(Lista_Carga, sheet_name=0)  # Primeira aba

        if coluna_chave not in df_lista.columns:
            print(f"Coluna '{coluna_chave}' não encontrada em {Lista_Carga}.")
            return set()

        codigos = (
            df_lista[coluna_chave]
            .dropna()
            .astype(str)
            .str.strip()
            .str.upper()
        )

        # Aplicar zfill(18) apenas para MATNR; para KUNNR e LIFNR usar zfill(10)
        if coluna_chave == "MATNR":
            codigos = codigos.str.zfill(18)
        elif coluna_chave in ["KUNNR", "LIFNR"]:
            codigos = codigos.str.zfill(10)

        return set(codigos.tolist())

    except Exception as e:
        print(f"Erro ao carregar '{Lista_Carga}': {e}")
        return set()

# Para Produto - # XXX AJUSTE AQUI XXX
codigos_produto = carregar_lista_carga("MATNR")
#codigos_cliente = carregar_lista_carga("KUNNR")
#codigos_fornecedor = carregar_lista_carga("LIFNR")


#############################################**** DE_PARA ****#############################################################################################
def buscar_valor_mapeado(coluna_destino, valor_original, df_de_para):
    """
    Versão que preserva zeros à esquerda
    """
    coluna_destino_upper = coluna_destino.upper()
    
    # Preservar formato original se for código
    valor_para_comparacao = str(valor_original).strip().upper() if valor_original is not None else ""
    
    # Verifica ALL primeiro
    mapeado_all = df_de_para[
        (df_de_para["COLUNA"].str.upper() == coluna_destino_upper) &
        (df_de_para["DE"].astype(str).str.upper() == "ALL")
    ]
    if not mapeado_all.empty:
        resultado = mapeado_all.iloc[0]["PARA"]
        return preservar_formato_original(resultado)

    # Verifica mapeamento direto
    mapeado = df_de_para[
        (df_de_para["COLUNA"].str.upper() == coluna_destino_upper) &
        (df_de_para["DE"].astype(str).str.upper() == valor_para_comparacao)
    ]
    if not mapeado.empty:
        resultado = mapeado.iloc[0]["PARA"]
        return preservar_formato_original(resultado)

    # Retorna valor original preservando formato
    return preservar_formato_original(valor_original)


def preencher_planilha_destino(df_origem, mappings, wb_destino, df_de_para):
    """
    Preenche os dados no workbook de destino conforme o mapeamento.
    VERSÃO CORRIGIDA: Preenche linha por linha, não coluna por coluna.
    """
    
    # Agrupar mappings por aba
    mappings_por_aba = {}
    for mapping in mappings:
        nome_aba = mapping['nome_aba']
        if nome_aba not in mappings_por_aba:
            mappings_por_aba[nome_aba] = []
        mappings_por_aba[nome_aba].append(mapping)
    
    # Processar cada aba
    for nome_aba, aba_mappings in mappings_por_aba.items():
        if nome_aba not in wb_destino.sheetnames:
            print(f"[ERRO] Aba '{nome_aba}' não encontrada. Pulando...")
            continue
            
        ws_destino = wb_destino[nome_aba]
        
        # Mapear colunas destino para seus índices
        colunas_destino = {}
        for mapping in aba_mappings:
            coluna_destino = mapping['coluna_destino']
            
            # Encontrar índice da coluna no destino
            col_dest_idx = None
            for col_idx, cell in enumerate(ws_destino[5], start=1):
                if cell.value and str(cell.value).strip().upper() == coluna_destino.upper():
                    col_dest_idx = col_idx
                    break
            
            if col_dest_idx is None:
                print(f"[ERRO] Coluna destino '{coluna_destino}' não encontrada na aba '{nome_aba}'.")
                continue
                
            # Resolver coluna de origem
            coluna_origem_real = resolver_coluna_origem(
                coluna_destino,
                df_origem,
                mapeamento_excecoes,
                colunas_equivalentes
            )
            
            colunas_destino[coluna_destino] = {
                'indice': col_dest_idx,
                'origem': coluna_origem_real
            }
        
        # AGORA SIM: Preencher linha por linha
        linha_destino = 9  # Linha inicial no destino
        
        for _, row_data in df_origem.iterrows():
            # Para cada linha da origem, preencher todas as colunas mapeadas
            for coluna_destino, config in colunas_destino.items():
                col_dest_idx = config['indice']
                coluna_origem_real = config['origem']
                
                # Se não tiver coluna real na origem, verificar ALL
                if not coluna_origem_real:
                    mapeado_all = df_de_para[
                        (df_de_para["COLUNA"].str.upper() == coluna_destino.upper()) &
                        (df_de_para["DE"].astype(str).str.upper() == "ALL")
                    ]
                    if not mapeado_all.empty:
                        valor_final = mapeado_all.iloc[0]["PARA"]
                    else:
                        valor_final = ""
                else:
                    # Obter valor da origem
                    valor_original = row_data.get(coluna_origem_real)
                    
                    # Se for NaN, tratar como string vazia
                    if pd.isna(valor_original):
                        valor_original = ""
                    
                    # Aplicar mapeamento De_Para
                    valor_final = buscar_valor_mapeado(coluna_destino, valor_original, df_de_para)
                
                # Preencher a célula
                #ws_destino.cell(row=linha_destino, column=col_dest_idx).value = valor_final
                # Preencher célula preservando zeros à esquerda
                cell_destino = ws_destino.cell(row=linha_destino, column=col_dest_idx)
                
                if eh_codigo_com_zeros(valor_final):
                    cell_destino.value = str(valor_final)
                    cell_destino.number_format = '@'  # Formato texto
                else:
                    cell_destino.value = valor_final                     
            
            # Avançar para próxima linha
            linha_destino += 1
        
        print(f"[INFO] Preenchimento linha por linha concluído na aba '{nome_aba}' - {len(df_origem)} linhas processadas.")
    
    print("[INFO] Preenchimento de todas as abas concluído (linha por linha).")


###############################################**** Preencher_Dados_nova ****#####################################################

# Mapeamento de exceções fixas
mapeamento_excecoes = {
    "TELNR_LONG_2": "TELF2",
    "TELNR_LONG_3": "TELF3",
    "ZWELS_01": "ZWELS",
    "ZTERM1": "ZTERM",
    "NAME_FIRST_P": "NAME1",
    "BPKIND": "KTOKK",
    "BP_ROLE": "RLTYP",
    "PRODUCT": "MATNR",
    "BISMT" : "MATNR"

}

# Colunas equivalentes por tipo
colunas_equivalentes = {
    "LIFNR": ["LIFNR", "PARTNER"],
    "KUNNR": ["KUNNR", "CUSTOMER"],
    "MATNR": ["MATNR", "PRODUCT"],
}  



########################################
from collections import defaultdict

from collections import defaultdict

def preencher_dados(tipo_execucao):
    """
    Preenche dados no template a partir do SQL Server ou planilha local,
    aplicando filtragem pela Lista_Carga e garantindo formatação exata para campos como TRAGR.
    """

    print(f"\n Iniciando preenchimento para template '{tipo_execucao.upper()}'")

    # Definir coluna-chave conforme tipo
    coluna_chave = coluna_chave_execucao.get(tipo_execucao)
    if not coluna_chave:
        print(f"[ERRO] Tipo de template '{tipo_execucao}' inválido.")
        return

    # Carregar lista de códigos (Lista_Carga)
    lista_carga = carregar_lista_carga(coluna_chave)
    print(f"[OK] {len(lista_carga)} códigos carregados para coluna '{coluna_chave}'.")

    # Abrir workbook de destino
    wb_destino = openpyxl.load_workbook(arquivo_destino)

    # Carregar planilha de mapeamento
    wb_mapeamento = openpyxl.load_workbook(planilha_mapeamento, data_only=True)
    ws_mapeamento = wb_mapeamento["Mapa"]

    # Carregar De_Para
    arquivo_de_para = 'f:/BackOffice_GERAL/Projeto S4 Hana/Onda 2/9. Saneamento/Asantos/25 . Saneamento/De_Para.xlsx'
    df_de_para = pd.read_excel(arquivo_de_para)

    if usar_sql_server:
        print("🔧 Modo: SQL Server")

        mapeamentos_por_tabela = defaultdict(list)
        for row in ws_mapeamento.iter_rows(min_row=2, values_only=True):
            nome_aba, planilha_origem, coluna_origem_destino = row
            mapeamentos_por_tabela[planilha_origem].append({
                'nome_aba': nome_aba,
                'coluna_origem_real': coluna_origem_destino,
                'coluna_destino': coluna_origem_destino
            })

        conn_str = (
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "SERVER=SPLCPVMSQLQA,23002;"
            "DATABASE=MDM_ONDA3;"
            "UID=usr_mdm;"
            "PWD=HSw0raUlpVcBC;"
            "TrustServerCertificate=yes;"
        )
        conn_str_escaped = urllib.parse.quote_plus(conn_str)
        engine = create_engine(f"mssql+pyodbc:///?odbc_connect={conn_str_escaped}")

        with engine.connect() as conn:
            for tabela_origem, mappings in mapeamentos_por_tabela.items():
                print(f"🔎 Consultando tabela '{tabela_origem}'...")

                # Montar filtro SQL (aplicando lista_carga)
                if lista_carga:
                    lista_str = "','".join(lista_carga)
                    query_sql = f"SELECT * FROM {tabela_origem} WHERE {coluna_chave} IN ('{lista_str}')"
                else:
                    query_sql = f"SELECT * FROM {tabela_origem}"

                try:
                    df_origem = pd.read_sql_query(query_sql, conn, dtype={coluna_chave: str})
                    print(f"✅ Consulta retornou {len(df_origem)} registros.")
                except Exception as e:
                    print(f"[ERRO] Falha ao consultar tabela '{tabela_origem}': {e}")
                    continue

                # Filtragem adicional em pandas (garantia)
                if lista_carga and coluna_chave in df_origem.columns:
                    df_origem = df_origem[df_origem[coluna_chave].isin(lista_carga)]

                # Preencher dados
                preencher_planilha_destino(df_origem, mappings, wb_destino, df_de_para)

    else:
        print("🔧 Modo: Planilha local")

        for row in ws_mapeamento.iter_rows(min_row=2, values_only=True):
            nome_aba, planilha_origem, coluna_origem_destino = row

            caminho_planilha_origem = f"Z:/00 Pastas de trabalho/Asantos/08 - Lista_Carga/Base_Excell/{planilha_origem}.xlsx"
            if not os.path.exists(caminho_planilha_origem):
                print(f"[ERRO] Planilha local '{caminho_planilha_origem}' não encontrada.")
                continue

            try:
                df_origem = pd.read_excel(caminho_planilha_origem, dtype=str)
                print(f"[OK] Planilha '{planilha_origem}' carregada com {len(df_origem)} registros.")

                # Filtrar pela lista de carga se existir
                if lista_carga and coluna_chave in df_origem.columns:
                    df_origem[coluna_chave] = df_origem[coluna_chave].astype(str).str.zfill(18)
                    df_origem = df_origem[df_origem[coluna_chave].isin(lista_carga)]
                    print(f"[OK] Filtrados {len(df_origem)} registros após aplicar Lista_Carga.")




                print(f"[OK] Planilha '{planilha_origem}' carregada com {len(df_origem)} registros.")
            except Exception as e:
                print(f"[ERRO] Falha ao carregar '{caminho_planilha_origem}': {e}")
                continue

            # Filtragem pela lista de carga
            if lista_carga and coluna_chave in df_origem.columns:
                df_origem = df_origem[df_origem[coluna_chave].isin(lista_carga)]

            mappings = [{
                'nome_aba': nome_aba,
                'coluna_origem_real': coluna_origem_destino,
                'coluna_destino': coluna_origem_destino
            }]

            preencher_planilha_destino(df_origem, mappings, wb_destino, df_de_para)

    # Salvar workbook de destino após preenchimento
    wb_destino.save(arquivo_destino_preenchido)
    print(f"[OK] Planilha preenchida e salva em: {arquivo_destino_preenchido}")

# ================== FUNÇÃO SALVAR LOGS ====================================================================================
def salvar_logs():
    wb_log = openpyxl.Workbook()
    ws_log = wb_log.active
    ws_log.title = "Log"
    for log in logs:
        ws_log.append(log)
    wb_log.save(arquivo_log)
    print(f"Logs salvos em: {arquivo_log}")

# ================== EXECUÇÃO ===============================================================================================
def reexibir_linha_5(workbook):
    for ws in workbook.worksheets:
        for cell in ws[5]:
            cell.font = openpyxl.styles.Font(bold=True)  # exemplo: deixa negrito


# Carregar workbook de destino
wb_destino = openpyxl.load_workbook(arquivo_destino)

# Carregar workbook de mapeamento
wb_mapeamento = openpyxl.load_workbook(planilha_mapeamento, data_only=True)
ws_mapeamento = wb_mapeamento.active

reexibir_linha_5(wb_destino)
preencher_dados(tipo_execucao)
salvar_logs()


###############################################**** FORMATO ****###############################################################
from openpyxl import load_workbook
from openpyxl.styles import numbers
import re
from tqdm import tqdm

def deve_preservar_zeros(valor):
    """
    Detecta dinamicamente se um valor deve preservar zeros à esquerda
    """
    if not isinstance(valor, str):
        return False
    
    valor_clean = str(valor).strip()
    
    # Deve ser só dígitos
    if not valor_clean.isdigit():
        return False
    
    # Deve começar com zero e ter mais de 1 dígito
    if valor_clean.startswith('0') and len(valor_clean) > 1:
        return True
    
    return False

# Carrega a planilha
wb = load_workbook(arquivo_destino_preenchido)

# Regex para identificar valores numéricos com vírgulas ou pontos
regex_numerico = re.compile(r"^[\d\.,]+$")
def converter_valor(celula):
    """
    Converte valores preservando códigos com zeros à esquerda
    """
    if celula.value is None:
        return
    
    valor_original = str(celula.value).strip()
    
    # Se já é número, não mexer
    if isinstance(celula.value, (int, float)):
        return
    
    # PRIORIDADE: Se é código com zeros, preservar como texto
    if eh_codigo_com_zeros(valor_original):
        celula.value = valor_original
        celula.number_format = '@'
        return
    
    # Se é só dígitos (sem zeros à esquerda), não mexer
    if valor_original.isdigit():
        return
    
    # Processar valores com vírgulas/pontos (números decimais)
    regex_numerico = re.compile(r"^[\d\.,]+$")
    if regex_numerico.match(valor_original):
        valor_convertido = valor_original.replace('.', '').replace(',', '.')
        try:
            numero = float(valor_convertido)
            celula.value = numero
            casas_decimais = len(valor_original.split(',')[1]) if ',' in valor_original else 0
            formato_excel = f'#,##0{"." + "0"*casas_decimais if casas_decimais else ""}'
            celula.number_format = formato_excel
        except ValueError:
            pass

# Calcular total de células para mostrar progresso
total_celulas = sum(len(linha) for aba in wb.worksheets for linha in aba.iter_rows())

# Processar com tqdm (barra de progresso)
with tqdm(total=total_celulas, desc="Processando células", ncols=100) as pbar:
    for aba in wb.worksheets:
        for linha in aba.iter_rows():
            for celula in linha:
                converter_valor(celula)
                pbar.update(1)

# Salvar o novo arquivo
wb.save(arquivo_destino_preenchido_LP)
print(f"\n Arquivo salvo como: {arquivo_destino_preenchido_LP}")

#############################################*** LIMPEZA ***###############################################

import openpyxl
from tqdm import tqdm


# Carregar o arquivo Excel
wb = openpyxl.load_workbook(arquivo_destino_preenchido_LP)

# Lista expandida de formatos que representam zero decimal
formatos_zero = [
    "0,000", "0,00", "0,0", "#,##0.000", "#,##0.00", "#,##0.0",  
    "#,##0", "#,###0", "#.##0,00", "#.##0,0", "#.##0", 
    "0.000", "0.00", "0.0", "00000000"
]

# Processar todas as abas
for aba in tqdm(wb.sheetnames, desc="Processando abas"):
    ws = wb[aba]

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value == 0:
                # Capturar o formato da célula e remover aspas se existirem
                formato = cell.number_format.replace('"', '').strip()
                
                # Se o formato da célula estiver na lista de zeros formatados, remover valor
                if any(f in formato for f in formatos_zero):
                    cell.value = None  # Apaga o valor sem alterar o formato
                # NOVO: Apagar somente strings com exatamente oito zeros
            elif isinstance(cell.value, str):
                valor_limpo = cell.value.strip()
                if valor_limpo == "00000000":
                    cell.value = None  # Apaga apenas "00000000"    

# Salvar mantendo o layout original
wb.save(arquivo_destino_preenchido_LP)

print("Processamento concluído. Arquivo salvo em:",arquivo_destino_preenchido_LP )

###############################################*** FINAL ***###############################################